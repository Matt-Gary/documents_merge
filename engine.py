"""
DRE Engine — handles all spreadsheet reading/writing logic.
No GUI dependencies; can be used standalone or tested independently.
"""

from __future__ import annotations
import re
from pathlib import Path
from datetime import datetime, date
import openpyxl
from openpyxl import load_workbook

# ─── Google Sheets target ────────────────────────────────────────────────────
GDRIVE_DRE_ID = "1nfETgiPN5pNrcIDDiklG5yptL6ltyE8kYOfUbjRl2zM"
CREDENTIALS_FILE = Path(__file__).parent / "spreadsheetexport-490117-f905b0cadf9e.json"
GSHEETS_SCOPES = ["https://www.googleapis.com/auth/spreadsheets"]


class ImportError(Exception):
    pass


# ─────────────────────────── SOURCE DETECTION ────────────────────────────────

def detect_source_type(filepath: Path) -> str:
    """
    Returns one of: 'cartao', 'caixa', 'bradesco', 'ai'
    PDFs and unrecognised XLSX files return 'ai' (require OpenAI analysis).
    """
    # PDFs always go through AI
    if Path(filepath).suffix.lower() == ".pdf":
        return "ai"

    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        sheet_names = [s.lower() for s in wb.sheetnames]
        first_sheet = wb.worksheets[0]
        rows = list(first_sheet.iter_rows(max_row=4, values_only=True))
        wb.close()
    except Exception:
        return "ai"

    # Cartão: sheet name contains 'nu_' or headers are Data/Valor/Identificador/Descrição
    for name in sheet_names:
        if name.startswith("nu_") or "pagveloz" in name or "nubank" in name:
            return "cartao"

    # Caixa PJ: sheet name contains 'extrato-pj' or headers Data/Histórico/Documento/Valor
    for name in sheet_names:
        if "extrato-pj" in name or "extrato_pj" in name or "caixa" in name:
            return "caixa"

    # Inspect headers
    all_cells = []
    for row in rows:
        for cell in row:
            if cell and isinstance(cell, str):
                all_cells.append(cell.lower().strip())

    cell_text = " ".join(all_cells)

    if "identificador" in cell_text and "descrição" in cell_text:
        return "cartao"

    if "histórico" in cell_text and "documento" in cell_text and "saldo" in cell_text:
        return "caixa"

    if "favorecido" in cell_text or "tipo" in cell_text and "saída" in cell_text.replace("í","i"):
        return "bradesco"

    if "consolidação" in cell_text or "consolidacao" in cell_text:
        return "bradesco"

    # Bradesco: sheet name
    for name in sheet_names:
        if "bradesco" in name or "consolidação" in name or "consolidacao" in name or "lançamentos" in name:
            return "bradesco"

    return "ai"


# ─────────────────────────── VALUE PARSING ───────────────────────────────────

def _parse_value(val) -> float | None:
    """Normalize a value from any source into a float."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        # Remove currency symbols and spaces, normalize decimal separator
        cleaned = val.strip().replace("R$", "").replace(" ", "")
        # Brazilian format: 1.234,56 → 1234.56
        if re.search(r"\d\.\d{3},\d{2}", cleaned):
            cleaned = cleaned.replace(".", "").replace(",", ".")
        else:
            cleaned = cleaned.replace(",", ".")
        try:
            return float(cleaned)
        except ValueError:
            return None
    return None


def _parse_date(val) -> date | None:
    """Normalize a date value."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y"):
            try:
                return datetime.strptime(val.strip(), fmt).date()
            except ValueError:
                continue
    return None


# ─────────────────────────── EXTRACTORS ──────────────────────────────────────

def _find_header_row(ws, keywords: list[str]) -> int | None:
    """Find the row index (0-based) where all keywords appear."""
    for i, row in enumerate(ws.iter_rows(max_row=10, values_only=True)):
        row_text = [str(c).lower().strip() if c else "" for c in row]
        row_joined = " ".join(row_text)
        if all(kw.lower() in row_joined for kw in keywords):
            return i
    return None


def _col_index(header_row: tuple, keyword: str) -> int | None:
    """Return the 0-based column index for the first cell matching keyword."""
    keyword_lower = keyword.lower()
    for i, cell in enumerate(header_row):
        if cell and keyword_lower in str(cell).lower():
            return i
    return None


def extract_cartao(filepath: Path, sheet_target: str) -> list[dict]:
    wb = load_workbook(filepath, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        raise ImportError("Arquivo do Cartão está vazio.")

    # Header: Data | Valor | Identificador | Descrição
    header = rows[0]
    col_data   = _col_index(header, "Data")
    col_valor  = _col_index(header, "Valor")
    col_descr  = _col_index(header, "Descrição")

    if None in (col_data, col_valor, col_descr):
        raise ImportError(f"Cabeçalhos esperados não encontrados no Cartão.\nEncontrado: {header}")

    result = []
    for row in rows[1:]:
        if not any(row):
            continue
        data  = _parse_date(row[col_data])
        valor = _parse_value(row[col_valor])
        desc  = str(row[col_descr]).strip() if row[col_descr] else ""

        if data is None or valor is None or not desc:
            continue

        target = "DESPESAS" if valor < 0 else "RECEITAS"

        result.append({
            "sheet":     target,
            "data":      data,
            "historico": desc,
            "valor":     valor,
        })
    return result


def extract_caixa(filepath: Path, sheet_target: str) -> list[dict]:
    wb = load_workbook(filepath, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # Caixa has account info on row 0; header on row 1
    header_idx = _find_header_row(ws if False else type("_", (), {"iter_rows": lambda *a, **k: iter(rows)})(),
                                   ["Data", "Histórico", "Valor"])
    # simpler: scan manually
    header_idx = None
    for i, row in enumerate(rows):
        row_text = " ".join(str(c) for c in row if c)
        if "Data" in row_text and "Histórico" in row_text and "Valor" in row_text:
            header_idx = i
            break

    if header_idx is None:
        raise ImportError("Cabeçalho não encontrado no arquivo da Caixa PJ.")

    header = rows[header_idx]
    col_data    = _col_index(header, "Data")
    col_hist    = _col_index(header, "Histórico")
    col_valor   = _col_index(header, "Valor")

    if None in (col_data, col_hist, col_valor):
        raise ImportError(f"Cabeçalhos esperados não encontrados na Caixa PJ.\nEncontrado: {header}")

    result = []
    for row in rows[header_idx + 1:]:
        if not any(row):
            continue
        data  = _parse_date(row[col_data])
        valor = _parse_value(row[col_valor])
        hist  = str(row[col_hist]).strip() if row[col_hist] else ""

        if data is None or valor is None or not hist:
            continue

        target = "DESPESAS" if valor < 0 else "RECEITAS"

        result.append({
            "sheet":     target,
            "data":      data,
            "historico": hist,
            "valor":     valor,
        })
    return result


def extract_bradesco(filepath: Path, sheet_target: str) -> list[dict]:
    wb = load_workbook(filepath, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # Find header row
    header_idx = None
    for i, row in enumerate(rows):
        row_text = " ".join(str(c) for c in row if c)
        if "Descrição" in row_text and "Valor" in row_text:
            header_idx = i
            break

    if header_idx is None:
        raise ImportError("Cabeçalho não encontrado no arquivo do Bradesco.")

    header = rows[header_idx]
    col_data    = 0  # first column is always date (unlabeled in Bradesco)
    col_descr   = _col_index(header, "Descrição")
    col_favor   = _col_index(header, "Favorecido")
    col_valor   = _col_index(header, "Valor")
    col_tipo    = _col_index(header, "Tipo")

    if None in (col_descr, col_valor):
        raise ImportError(f"Cabeçalhos esperados não encontrados no Bradesco.\nEncontrado: {header}")

    auto_mode = sheet_target == "Auto (Tipo)"

    result = []
    for row in rows[header_idx + 1:]:
        if not any(row):
            continue
        data  = _parse_date(row[col_data])
        valor = _parse_value(row[col_valor])
        descr = str(row[col_descr]).strip() if row[col_descr] else ""
        favor = str(row[col_favor]).strip() if col_favor is not None and row[col_favor] else ""
        tipo  = str(row[col_tipo]).strip() if col_tipo is not None and row[col_tipo] else ""

        if data is None or valor is None or not descr:
            continue

        # Build historico: "Descrição - Favorecido/Origem"
        historico = f"{descr} - {favor}" if favor and favor not in ("None", "") else descr

        # Determine destination sheet
        if auto_mode:
            # "Saída (-)" → DESPESAS, "Entrada (+)" → RECEITAS
            if "saída" in tipo.lower() or "saida" in tipo.lower():
                target = "DESPESAS"
                valor = abs(valor) * -1  # ensure negative for expenses
            elif "entrada" in tipo.lower():
                target = "RECEITAS"
                valor = abs(valor)
            else:
                target = "DESPESAS"  # fallback
        else:
            target = sheet_target

        result.append({
            "sheet":     target,
            "data":      data,
            "historico": historico,
            "valor":     valor,
        })
    return result


# ─────────────────────────── MAIN ENGINE ─────────────────────────────────────

class DREEngine:

    def extract_rows(
        self,
        filepath: Path,
        source_type: str,
        sheet_target: str,
        api_key: str | None = None,
    ) -> list[dict]:
        """Extract rows from a source file and return normalized dicts."""
        if source_type == "cartao":
            rows = extract_cartao(filepath, sheet_target)
        elif source_type == "caixa":
            rows = extract_caixa(filepath, sheet_target)
        elif source_type == "bradesco":
            rows = extract_bradesco(filepath, sheet_target)
        elif source_type == "ai":
            if not api_key:
                raise ImportError(
                    f"Arquivo '{filepath.name}' requer análise de IA.\n"
                    "Informe a OpenAI API Key no campo acima."
                )
            from ai_mapper import extract_with_ai
            rows = extract_with_ai(filepath, api_key)
        else:
            raise ImportError(
                f"Tipo de arquivo não reconhecido: {filepath.name}\n"
                "Formatos suportados: Cartão PagVeloz, Caixa PJ, Bradesco, PDF (via IA)."
            )
        # Stamp every row with the source filename
        for r in rows:
            r["source"] = filepath.name
        return rows

    def write_to_dre(self, rows: list[dict]) -> int:
        """
        Append rows to the DRE Google Sheets spreadsheet.
        Uses service account credentials from credentials.json.
        Returns number of rows inserted.
        """
        try:
            import gspread
        except ImportError:
            raise ImportError(
                "Pacotes necessários não instalados.\n"
                "Execute: pip install gspread google-auth"
            )

        if not CREDENTIALS_FILE.exists():
            raise ImportError(
                f"Arquivo de credenciais não encontrado:\n{CREDENTIALS_FILE}\n\n"
                "Coloque o arquivo credentials.json na pasta do programa."
            )

        try:
            gc = gspread.service_account(filename=str(CREDENTIALS_FILE))
        except Exception as e:
            import traceback
            raise ImportError(
                f"Falha na autenticação com a conta de serviço.\n\n{traceback.format_exc()}"
            )

        try:
            sh = gc.open_by_key(GDRIVE_DRE_ID)
        except Exception as e:
            import traceback
            raise ImportError(
                f"Não foi possível abrir a planilha Google Sheets.\n"
                f"Certifique-se de que ela foi compartilhada com a conta de serviço.\n\n"
                f"{traceback.format_exc()}"
            )

        inserted = 0
        for sheet_name in ("DESPESAS", "RECEITAS"):
            sheet_rows = [r for r in rows if r["sheet"] == sheet_name]
            if not sheet_rows:
                continue

            try:
                ws = sh.worksheet(sheet_name)
            except Exception:
                raise ImportError(f"Aba '{sheet_name}' não encontrada na planilha DRE.")

            # Find last non-empty row in column A specifically
            col_a = ws.col_values(1)
            last_row = 1
            for idx in range(len(col_a) - 1, -1, -1):
                if col_a[idx] and str(col_a[idx]).strip():
                    last_row = idx + 1  # 1-based
                    break

            # Expand sheet if needed so we don't exceed grid limits
            rows_needed = len(sheet_rows)
            current_max = ws.row_count
            first_target_row = last_row + 1
            target_last = last_row + rows_needed
            if target_last > current_max:
                ws.add_rows(target_last - current_max + 50)

            # Build all rows as a 2D list and write in ONE API call
            values = []
            for r in sheet_rows:
                date_str = r["data"].strftime("%d/%m/%Y") if hasattr(r["data"], "strftime") else str(r["data"])
                values.append([date_str, r["historico"], "", r["valor"]])

            ws.update(
                range_name=f"A{first_target_row}:D{target_last}",
                values=values,
                value_input_option="USER_ENTERED",
            )
            inserted += len(values)

        return inserted